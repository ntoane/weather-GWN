import Utils.gwnUtils as utils
import Utils.metrics as metrics
import numpy as np
import openpyxl 


def eval(stations, args):
    """
     Calculates the GWN model's performance on the test set across all forecasting horizons[1, 3, 6, 12, 24, 48] for each
     individual station. The predictions are read from the results file for each split of the walk-forward validation
     method. The predictions from each split are appended into one long list of predictions. The targets are pulled from
     targets file in the GWN results directory. The MSE, MAE, RMSE, and SMAPE metrics are then calculated and written
     to the metric files.

     Parameters:
         stations - List of the weather stations.
         args - Parser of parameter arguments.
     """
    num_splits = 26
    num_stations = 27

    path = "Results/GWN/Metrics/Stations_matrics.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active

    for station in range(num_stations):
        # Start writing first station at row 3, everytime we add 3 to station variable
        row = 3 + station
        # Write station number from A3(i.e, on each Arow)
        station_cell = sheet_obj.cell(row = row, column = 1)
        station_cell.value = station + 1

        horizon_counter = 2

        for horizon in [1, 3, 6, 12, 24, 48]:

            pred = []
            real = []

            for split in range(num_splits):
                resultsFile = 'Results/GWN/' + str(horizon) + ' Hour Forecast/Predictions/outputs_' + str(split) + \
                              '.pkl'
                targetsFile = 'Results/GWN/' + str(horizon) + ' Hour Forecast/Targets/targets_' + str(split) + '.pkl'
                yhat = utils.load_pickle(resultsFile)
                target = utils.load_pickle(targetsFile)
                pred.extend(np.array(yhat).flatten())
                real.extend(np.array(target).flatten())

            pred = np.array(pred).reshape((int(len(real) / (args.n_stations * args.seq_length)), args.n_stations,
                                           args.seq_length))
            real = np.array(real).reshape((int(len(real) / (args.n_stations * args.seq_length)), args.n_stations,
                                           args.seq_length))

            metricFile = 'Results/GWN/Metrics/' + stations[station] + '/metrics_' + str(horizon)
            file = open(metricFile, 'w')

            preds = pred[:, station, :]
            real_values = real[:, station, :]

            rmse = metrics.rmse(real_values, preds)
            mse = metrics.mse(real_values, preds)
            mae = metrics.mae(real_values, preds)
            smape = metrics.smape(real_values, preds)

            print('MSE: {0} for station {1} forecasting {2} hours ahead'.format(mse, station+1, horizon))
            print('RMSE: {0} for station {1} forecasting {2} hours ahead'.format(rmse, station+1, horizon))
            print('MAE: {0} for station {1} forecasting {2} hours ahead'.format(mae, station+1, horizon))
            print('SMAPE: {0} for station {1} forecasting {2} hours ahead'.format(smape, station+1, horizon))
            print(' ')

            file.write('MSE: ' + str(mse) + '\n')
            file.write('RMSE: ' + str(rmse) + '\n')
            file.write('MAE: ' + str(mae) + '\n')
            file.write('SMAPE: ' + str(smape) + '\n')

            file.close()

            #Writing to Excel file
            # Columns starts from B, so first MSE is written at B3, then F3, J3, N3, R3, V3
            rmse_cell = sheet_obj.cell(row = row, column = horizon_counter)
            mse_cell = sheet_obj.cell(row = row, column = horizon_counter + 1)
            mae_cell = sheet_obj.cell(row = row, column = horizon_counter + 2)
            smape_cell = sheet_obj.cell(row = row, column = horizon_counter + 3)

            # Write RMSE, MSE, MAE, SMAPE values to excel sheet
            rmse_cell.value = rmse
            mse_cell.value = mse
            mae_cell.value = mae
            smape_cell.value = smape

            # Increment horizon counter by 4 to write next horizon's metrics
            horizon_counter += 4

    # Save the excel file
    wb_obj.save(path)
